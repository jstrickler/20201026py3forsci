#!/usr/bin/env python
import openpyxl as px

fruits = ["pomegranate", "cherry", "apricot", "date", "apple", "lemon", "kiwi",
          "orange", "lime", "watermelon", "guava", "papaya", "fig", "pear", "banana",
          "tamarind", "persimmon", "elderberry", "peach", "blueberry", "lychee",
          "grape"]

extra = [(len(f), f.upper()) for f in fruits]
print(extra)

wb = px.Workbook()

ws = wb.active

ws.title = 'fruits'

ws_new = wb.create_sheet("New stuff")

for i, fruit in enumerate(fruits, 1):
    ws.cell(row=i, column=1).value = fruit
    length, upper_version = extra.pop(0)  # removes from list
    ws.cell(row=i, column=2).value = length
    ws.cell(row=i, column=3).value = upper_version

    ws_new.cell(row=i, column=1).value = upper_version

ws2 = wb.copy_worksheet(ws)

ws2["A1"] = "COPY!!"

wb.save('fruits.xlsx')
